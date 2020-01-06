import sys
import os
import random

os.system("pip3 install --user -r requirements.txt -q")
import openpyxl


class Participant:
    def __init__(self, id, score=None):
        self.id = id
        self.score = score
        self.partner = None
        self.payoff = 0

    def set_partner(self, partner):
        self.partner = partner

    def set_payoff(self, payoff):
        self.payoff = payoff

    def get_attributes(self):
        if self.score is not None:
            return [self.id, self.score, self.partner.id, self.payoff]
        else:
            return [self.id, self.partner.id, self.payoff]


def save_header(col_ids, labels, ws):
    for i in range(0, len(col_ids)):
        ws.cell(row=1, column=col_ids[i]).value = labels[i]


def save_participant_info(participant_list, col_ids, ws):
    starting_row = 2
    for p in participant_list:  # Save values to the excel file
        attributes = p.get_attributes()
        for i in range(0, len(attributes)):
            ws.cell(row=starting_row, column=col_ids[i]).value = attributes[i]
        starting_row += 1


def compute_scrambling_matching(file, ws, low_win, high_win):
    info = file.readlines()[1:]
    participant_list = [Participant(int(i)) for i in info]

    if len(participant_list) % 2 != 0:  # If it's odd, one person will be unmatched
        print("The number of participants is currently odd, we will not match the last participant")
        participant_list = participant_list[:-1]

    half_list = int(len(participant_list) / 2)

    for i in range(0, half_list):
        p = participant_list[i]  # We match the people from the first half
        p.partner = participant_list[i + half_list]  # With the people from the second half
        p.partner.partner = p

    id_col = 1
    part_col = 2
    payoff_col = 3
    save_header([id_col, part_col, payoff_col], ["participant_id", "partner_id", "payoff"], ws)

    for i in range(0, half_list):
        p = participant_list[i]  # For each person from the first half
        if random.randint(0, 1):  # Roll a die
            p.payoff = high_win  # Set their win and their partners
            p.partner.payoff = low_win
        else:
            p.payoff = low_win
            p.partner.payoff = high_win

    save_participant_info(participant_list, [id_col, part_col, payoff_col], ws)


def compute_click_matching(file, ws, low_win, high_win):
    info = file.readlines()[1:]
    participant_list = []
    for i in info:
        splitted = i.split(',')
        participant_list.append(Participant(int(splitted[0]), int(splitted[1])))

    id_col = 1
    score_col = 2
    part_col = 3
    payoff_col = 4

    save_header([id_col, score_col, part_col, payoff_col], ["participant_id", "score", "partner_id", "payoff"], ws)

    unmatched_list = participant_list.copy()  # Copy the list of participants
    unmatched_list.sort(key=lambda x: x.score)  # Order it by score
    # print([(p.id, p.score) for p in unmatched_list])
    while len(unmatched_list) > 1:  # If there is more than one participant
        p1 = unmatched_list.pop(0)  # Collect the first participant of the list
        # Find the first one with different score
        p2_index = next((i for i, x in enumerate(unmatched_list) if x.score != p1.score), None)
        if p2_index is None:  # If it doesn't exist
            p2 = unmatched_list[0]  # Choose the next one anyway
        else:
            p2 = unmatched_list[p2_index]  # Or choose the next one with different score
        unmatched_list.remove(p2)  # Remove the second participant from the list
        p1.partner = p2  # Set the partners
        p2.partner = p1
        if p1.score > p2.score:  # If the score of p1 is better
            p1.payoff = high_win  # Give him the high payoff
            p2.payoff = low_win
        else:
            p1.payoff = low_win  # Otherwise the low payoff
            p2.payoff = high_win

    if len(unmatched_list) == 1:  # If there is one unmatched person
        participant_list.remove(unmatched_list[0])  # Remove them from the list
        print("One participant was not matched because the number of participants was odd.")

    save_participant_info(participant_list, [id_col, score_col, part_col, payoff_col], ws)


if __name__ == "__main__":

    ass_type = int(sys.argv[1])
    csv_path = sys.argv[2]
    excel_path_name = sys.argv[3]

    ff = open(csv_path)
    if os.path.isfile(excel_path_name):
        wb = openpyxl.load_workbook(excel_path_name)
    else:
        wb = openpyxl.Workbook()

    ws = wb.create_sheet('Ass' + str(ass_type))

    if ass_type == 1 or ass_type == 2:
        low_win = 0
        high_win = 6
        compute_scrambling_matching(ff, ws, low_win, high_win)
    else:
        low_win = 0
        high_win = 6
        compute_click_matching(ff, ws, low_win, high_win)

    wb.save(excel_path_name)
